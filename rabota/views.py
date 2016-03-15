#-*- coding:utf-8 -*-
from django.shortcuts import render_to_response, RequestContext
from django.http import HttpResponseRedirect, HttpResponse, Http404
from rabota.forms import ContactForm, settings_cam, view_cam, raschet_form,Name_seria,Seria
from django.contrib.auth.decorators import login_required
from sites.models import navimenu
from rabota.models import cam,event,otchet_po_day,otchet_po_hears,serega,Seria_name
from qsstats import QuerySetStats
from django.core.mail import send_mail
from django.forms.models import inlineformset_factory

import datetime
import os
import math
import time

@login_required
def rabota(request):
    autuser = u'Вы авторизованные '

    if request.method == 'POST' and request.POST.has_key('submit'):
        form = ContactForm(request.POST, request.FILES)
        if form.is_valid():
            cd = form.cleaned_data
            ishod=cd['ishod'].read()
            shablon = request.POST['shablon']
            ishod = ishod.lower()
            shablon = shablon.lower()
            ishod_a = ishod.split(' ')
            shablon_a = shablon.split(' ')
            shablon_a = obrab(shablon_a)
            ishod_a = obrab(ishod_a)
            v = dictori()
            vari = request.POST['variant']
            vari1 = request.POST['variant1']
            vari2 = request.POST['variant2']
            result = []
            c = None
            wr = None
            a = None
            b = None
            d = None
            wb = None
            rr = None
            if vari == "1" and vari1 == "1":
                    a = progonka(ishod_a,1,v)
                    b = progonka(shablon_a,1,v)
                    result = vozrat(poln_pifag(2,a,b),ishod_a)
                    zapis11(vozrat1(poln_pifag(2,a,b),ishod_a))
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
            if vari == "1" and vari1 == "2":
                    a = progonka(ishod_a,1,v)
                    b = progonka(shablon_a,1,v)
                    result = vozrat(poln_pifag(1,a,b),ishod_a)
                    zapis11(vozrat1(poln_pifag(1,a,b),ishod_a))
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
            if vari == "2" and vari1 == "1":
                    a = progonka(ishod_a,2,v)
                    b = progonka(shablon_a,2,v)
                    result = vozrat(poln_pifag(2,a,b),ishod_a)
                    zapis(vozrat1(poln_pifag(2,a,b),ishod_a))
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
            if vari == "2" and vari1 == "2":
                    a = progonka(ishod_a,2,v)
                    b = progonka(shablon_a,2,v)
                    result = vozrat(poln_pifag(1,a,b),ishod_a)
                    zapis11(vozrat1(poln_pifag(1,a,b),ishod_a))
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
            if vari == "1" and vari1 == "3":
                    w = progonka(shablon_a,1,v)
                    qw = format(ishod_a,shablon_a)
                    c = progonka(qw,1,v)
                    wr = sun_pol(w)
                    we = poln_pifag(1,c,wr)
                    result = vozrat_pol(we,shablon_a,ishod_a)
                    zapis11(vozrat_pol1(we,shablon_a,ishod_a))
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
            if vari == "2" and vari1 == "3":
                    w = progonka(shablon_a,2,v)
                    qw = format(ishod_a,shablon_a)
                    c = progonka(qw,2,v)
                    wr = sun_pol(w)
                    we = poln_pifag(2,c,wr)
                    result = vozrat_pol(we,shablon_a,ishod_a)
                    zapis11(vozrat_pol1(we,shablon_a,ishod_a))
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
            if vari1 == "4":
                    b = progonka(shablon_a,1,v)
                    r = progonka(shablon_a,2,v)
                    qw = format(ishod_a,shablon_a)
                    c = progonka(qw,1,v)
                    d = progonka(qw,2,v)
                    wr = sun_pol(b)
                    wb = sun_pol(r)
                    we = poln_pifag(1,c,wr)
                    wq = poln_pifag(1,d,wb)
                    result = vozrat_pol(polrealpifag(we,wq,wb,wr),shablon_a,ishod_a)
                    zapis11(vozrat_pol1(polrealpifag(we,wq,wb,wr),shablon_a,ishod_a))
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
                    zapis1(' '.join(qw))
            if vari1 == "5":
                    b = progonka(ishod_a,1,v)
                    r = progonka(ishod_a,2,v)
                    wr = sun_pol(progonka(shablon_a,1,v))
                    wb = sun_pol(progonka(shablon_a,2,v))
                    result = Poiskporealpifag(b,r,wr,wb,ishod_a)
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
                    zapis11(result)
            if vari1 == "6":
                    q = list(ishod)
                    w = list(shablon)
                    q = obrab(q)
                    w = obrab(w)
                    b = progonka(q,2,v)
                    r = progonka(w,2,v)
                    result = pizdeckaknado(b,r,q)
                    result_a = result.split('.')
                    result = ' '.join(chisla(result_a,v))
                    zapis11(result)
            if vari == "3" and vari1 == "7" and vari2 == "1":
                    q = list(ishod)
                    q = obrab(q)
                    b = sun_pol(progonka(q,2,v))
                    a = sun_pol(progonka(q,1,v))
                    c = podschet(q)
                    result = ishod
            if vari == "3" and vari1 == "7" and vari2 == "2":
                q = list(ishod)
                q = obrab(q)
                b = sun_pol(progonka(q,2,v))
                a = sun_pol(progonka(q,1,v))
                c = podschet(q)
                result = ishod
                n = 0
                k = 0
                result_v,re = [],[]
                while n<len(result):
                    if result[n] == '.' or result[n] == ':' or result[n] == '?' or result[n] == '!':
                        res = result[k:n+1]
                        re.append(podschet(res))
                        k = n+1
                        pif = sun_pol(progonka(list(res),2,v))
                        rel = sun_pol(progonka(list(res),1,v))
                        chisl = podschet(res)
                        res = res+'%s%s' % (pif,rel)
                        res = res+'[%s]' % chisl
                        result_v.append(res)
                        n+=1
                    n+=1
                result_v.append(result[k:n])
                if result_v == []:
                    result = ishod
                else:
                    result = ''.join(result_v)
                    n=0
                    rr = '0'
                    for mn in re:
                        n+=mn
                        rr = rr +'+'+'%s' % mn
                    rr = rr +'='+ '%s' %n
            return render_to_response('rabota.html',{'result':unicode(result.decode('cp1251')),'a':a, 'b':b, 'c':c, 'wr':wr, 'd':d, 'wb':wb,'rr':rr, 'autuser':autuser},RequestContext(request))

    elif request.method == 'POST' and request.POST.has_key('button'):
        return HttpResponseRedirect('/statics/test.txt')
    elif request.method == 'POST' and request.POST.has_key('logout'):
        return HttpResponseRedirect('/accounts/logout/')
    else:
        form = ContactForm(initial={'shablon':u'Я люблю этот сайт'})
    return render_to_response('rabota.html',{'form':form, 'autuser':autuser},RequestContext(request))

class dictori:
# Инициализация словаря
    def __init__(self):
        self.a = []
        self.a.append(dict.fromkeys(['a','j','s'],1))
        self.a.append(dict.fromkeys(['b','k','t'],2))
        self.a.append(dict.fromkeys(['c','l','u'],3))
        self.a.append(dict.fromkeys(['d','m','v'],4))
        self.a.append(dict.fromkeys(['e','n','w'],5))
        self.a.append(dict.fromkeys(['f','o','x'],6))
        self.a.append(dict.fromkeys(['g','p','y'],7))
        self.a.append(dict.fromkeys(['h','q','z'],8))
        self.a.append(dict.fromkeys(['i','r'],9))
        self.a.append(dict.fromkeys(['.','/',':',';','"','`','-','_',',','1','2','3','4','5','6','7','8','9','0'],0))
        self.b = []
        self.b.append(dict([('a',1),('b',2),('c',3),('d',4),('e',5),('f',6),('g',7),('h',8),('i',9),('j',10),('k',11),('l',12),('m',13),('n',14),('o',15),('p',16),('q',17),('r',18),('s',19),('t',20),('u',21),('v',22),('w',23),('x',24),('y',25),('z',26),('/',0),('.',0),('/',0),(':',0),('"',0),('`',0),('-',0),('=',0),('_',0)]))
#Функция перефода букв в пифагор и обычную (пифагор vol2=1, j, обычный vol2 любой)
    def pifagor1(self, vol1, vol2):
        if vol2 == 1:
            self.zn = self.a
        else:
            self.zn = self.b
        n=0
        while n<len(self.zn):
            if self.zn[n].get(vol1):
                return self.zn[n].get(vol1)
            n+=1
        return self.a[9].get('-')
#Функция преоброзования списка по словам

def obrab(vol):
    vq=[]
    qv = []
    for n in vol:
        resul = n.split("\n")
        vq.append(resul)
    for ve in vq:
        for n in ve:
            if n =='':
                continue
            else:
                qv.append(n)
    return qv

def zapis11(vol):
    f = open('statics/test.txt','w')
    f.write(vol)
    f.close()
    return vol
def zapis1(vol):
    f = open('statics/test1.txt','w')
    f.write(vol)
    f.close()
    return vol

def format(vol1, vol2):
    m = 0
    n = 0
    u =[]
    y=[]
    while n<len(vol1):
        while m<len(vol2):
            u.append(vol1[n+m])
            m+=1
            if n+m == len(vol1):
                break
        m=0
        strok = ' '.join(u)
        y.append(strok)
        u=[]
        n+=1
    w=''
    e=[]
    for n in y:
        r = n.split(' ')
        for q in r:
           w = w+q
        e.append(w)
        w=''
    return e

# Функция сумирования чисел букв и возрощает список  = сумме числел букв в слове
def progonka(f,znach,v):
    a=0
    result = []
    result1 = 0
    result2 = []
    result3 = []
    for n in f:
       while a<len(n):
           result.append(v.pifagor1(n[a],znach))
           result1 = result1+int(v.pifagor1(n[a],znach))
           a+=1
       a=0
       result2.append(result1)
       result1 = 0
    return result2
# Поиск значений по шаблону vol = 1 полный поиск по шаблону, любое число поиск по словам. vol1 - исходный список vol2 - шаблон
def poln_pifag(vol,vol1,vol2):
    q=[]
    p = 0
    i=0
    for n in vol1:
        for m in vol2:
            if vol == 1:
                if n == m and n == vol2[i]:
                    q.append(dict([(n,p)]))
                    i+=1
                if i==len(vol2):
                    i=0
            else:
                if m == n:
                    q.append(dict([(n,p)]))
        p+=1
    return q
# функция форщает исходный масив с найдеными фрагментами необходми значения из функции poln_pifag
def vozrat(vol,vol1):
    for n in vol:
        for key,ind in n.items():
            v = vol1[ind].upper()
            vol1[ind] = v
    konec = ' '.join(vol1)
    return konec
def vozrat1(vol,vol1):
    qw = []
    for n in vol:
        for key,ind in n.items():
            v = vol1[ind].upper()
            qw.append(v)
    konec = ' '.join(qw)
    return konec

# Функция возрашает сумму всех чисел в предложении
def sun_pol(vol):
    rw=0
    wr=[]
    for n in vol:
        rw = rw+int(n)
    wr.append(rw)
    return wr

# Функция возрашает строку исходника.
def vozrat_pol(vol,vol1,vol2):
    for n in vol:
            for key,ind in n.items():
                t=0
                while t<len(vol1):
                    v = vol2[ind+t].upper()
                    vol2[ind+t] = v
                    t+=1
    konec = ' '.join(vol2)
    return konec

def vozrat_pol1(vol,vol1,vol2):
    df = []
    for n in vol:
            for key,ind in n.items():
                t=0
                while t<len(vol1):
                    v = vol2[ind+t].upper()
                    df.append(v)
                    t+=1
    konec = ' '.join(df)
    return konec
# как надо вариант 4
def polrealpifag(vol,vol1,vol2,vol3):
    re = []
    for n in vol:
        for r,t in n.items():
            for m in vol1:
                for rm,tm in m.items():
                    if rm == vol2[0] and tm == t and r == vol3[0]:
                        re.append(dict([(r,t)]))
    return re

def chisla(vol,vol2):
    iss = []
    for m in vol:
        m=m+'.'
        iss.append(m)
    nn=0
    vv = []
    cc = []
    while nn<len(vol):
        if len(iss[nn]) < 8:
            cc.append(iss[nn])
            nn+=1
        else:
            vv = iss[nn]+str(sun_pol(progonka(iss[nn],1,vol2)))+str(sun_pol(progonka(iss[nn],2,vol2)))
            cc.append(vv)
            nn+=1
    return cc
#Как надо вариант 5
def Poiskporealpifag(vol,vol1,vol2,vol3,vol4):
    m=0
    n=0
    ch = 0
    ch1 = 0
    vrem = []
    vrem1 = []
    while m<len(vol):
        ch = ch + vol[m]
        ch1 = ch1 + vol1[m]
        if ch == vol2[0] and ch1 == vol3[0]:
            vrem.append(n)
            vrem1.append(m+1)
            ch = 0
            ch1 = 0
            n+=1
            m = n
        if ch > vol2[0] or ch1 > vol3[0]:

            ch = 0
            ch1 = 0
            n+=1
            m = n
        else:
            m+=1
    m = 0
    while m<len(vrem):
        d = vrem1[m]-vrem[m]
        m1 = 0
        while m1<d:
            q = vol4[vrem[m]+m1].upper()
            vol4[vrem[m]+m1] = q
            m1+=1
        m+=1
    resulkaknado = ' '.join(vol4)
    return resulkaknado

def pizdeckaknado(vol1,vol2,vol3):
    n =0
    b = []
    x = []
    c = 0
    while n<len(vol1):
        m = 0
        while m<len(vol2):
            if n >= len(vol1)-len(vol2):
                break
            if vol2[m]== vol1[n+m]:
                b.append(vol1[n+m])
                if m == len(vol2)-1 and len(b)-1 == m:
                    c = 1
                m+=1
                if vol1[n+m] == 0:
                    n+=1
            elif c == 1:
                x.append(b)
                c = 0
                for t in xrange(n-2,n+len(vol2)-1):
                    vb = vol3[t].upper()
                    vol3[t] = vb
            else:
                b = []
                m+=1
        n+=1
        result = ''.join(vol3)
    return result

def podschet(vol):
    n=0
    c = 0
    m = ["'",'"',' ', '\r', '\n', '\x97','\x94','\x92','\x91','\x93','\t',')','(','.',',','!','?','`',':',';','1','2','3','4','5','6','7','8','9','0','*','=','+','-','/','|','^','@','[',']','{','}','!','?','&']
    while n<len(vol):
        t = 0
        y = False
        while t<len(m):
            if vol[n] == m[t]:
                  t+=1
                  y = True
                  break
            else:
                  t+=1
        if  y:
             n+=1
             continue
        else:
             c+=1
             n+=1
    return c

#Новый проект для видео наблюдения

def zapis(request):
    return render_to_response('zapis.html',{'request':request},RequestContext(request))

def process_zapisi(request):
    basedir = "e:\\ftp\detektor\kassa"
    b = os.listdir(basedir)
    file_names = []
    for n in b:
        p = n.split('_')
        if p[1] == 'MD.log':
            continue
        file_names.append(p[0])
        h = basedir+"\\"+n
        times = os.path.getmtime(h)
        filetime = datetime.datetime.strptime(time.ctime(times),'%a %b %d %H:%M:%S %Y')
        realtime = datetime.datetime.strptime(time.ctime(time.time()),'%a %b %d %H:%M:%S %Y')
        raznica = realtime-filetime
        time_nacalo = datetime.datetime.strptime("10:00:00",'%H:%M:%S')
        time_konec = datetime.datetime.strptime("20:00:00",'%H:%M:%S')
        if raznica > datetime.timedelta(hours=1) and time_nacalo.hour > 10 and time_konec.hour < 20:
            for nn in file_names[len(file_names)-1:]:
                kk = cam.objects.get(file_name = nn )
            message = u"Нет данных по камере %s более часа" % kk.name
            send_mail('Отвалилась камера',message, 'yatcina@mebtorg.ru',['yatcina@mebtorg.ru','kirill@mebtorg.ru'])
    if file_names == []:
        error = u'В папке нет файлов'
    else:
        error = u'Все заверщилось без ощибок'
        n=0
        c = inicialize()
        while n<len(file_names):
            l = cam.objects.get(file_name = file_names[n])
            url_file = 'e:\\ftp\detektor\kassa\%s_MD.syslog' % (l.file_name)
            result = c.loadfile(url_file)
            vhod_detektor2 = 'event'+ ' ' + l.detektor3
            vihod_detektor2 = 'event'+' ' + l.detektor4
            vhod_detektor1 = 'event'+' ' + l.detektor1
            vihod_detektor1 = 'event'+ ' ' + l.detektor2
            ishod_arry = c.baza(result,vhod_detektor2,vihod_detektor2,vhod_detektor1,vihod_detektor1)
            c.zapis_v_basu(ishod_arry,l)
            n+=1
    p=u'OK'
    return render_to_response('thanks.html',{'c':p,'error':error},RequestContext(request))


@login_required
def settings_camera(request,vol,vol1):
    q = 1
    mm =[]
    while q<4:
        mm.append(q)
        q+=1
    if cam.objects.filter(id = int(vol1)):
        kam_real = cam.objects.get(id = int(vol1))
    else:
        kam_real = cam.objects.get(id = 1)
    if int(vol) >= 01 and int(vol) <= 04:
        kam = cam.objects.all().order_by()[(int(vol)-1)*20:(int(vol)-1)*20+20]
        nas_kam = kam_real.name
        k = True
    else:
        nas_kam = u"камеры нет"
        k = True
        kam = None
        vrem = None
    if int(vol1) == int(kam_real.id) and int(vol) >= 01 and int(vol) <= 04:
        k = True
        vrem = cam.objects.get(id = int(vol1))
    else:
        vrem =None

    if request.method == 'POST':
        form = settings_cam(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            if int(vol1) == 0 and int(vol) == 00 or int(vol) >=4:
                m1 = cam.objects.create(name = cd['name'], adress = cd['adress'], detektor1 = cd['detektor1'], detektor2 = cd['detektor2'], detektor3 = cd['detektor3'], detektor4 = cd['detektor4'], file_name = cd['file_name'], ip_adress = cd['ip_adress'])
                m1.save()
            if int(vol1) == int(kam_real.id) and int(vol) >= 01 and int(vol) <= 04:
                vrem = cam.objects.get(id = int(vol1))
                vrem.name, vrem.adress = (cd['name'], cd['adress'])
                vrem.file_name, vrem.detektor4, vrem.detektor3 = (cd['file_name'],cd['detektor4'],cd['detektor3'])
                vrem.detektor1,vrem.detektor2 = (cd['detektor1'],cd['detektor2'])
                vrem.ip_adress = cd['ip_adress']
                vrem.save()
            return render_to_response('form_settings_cam.html',{'form':form, 'kam':kam, 'vol':vol,'vol1':vol}, RequestContext(request))
    else:
        form = settings_cam()
    return render_to_response('form_settings_cam.html',{'form':form, 'kam':kam,'vrem':vrem, 'mm':mm, 'k':k, 'nas_kam':nas_kam, 'vol':vol},RequestContext(request))


@login_required()
def shitalka(request):
        global p, data_time_start, data_time_start1, data_time_end, time_interval, variant
        if request.method == 'POST'  and request.POST.has_key('Otchet'):
            form1 = view_cam(request.POST)
            if form1.is_valid():
                cd = form1.cleaned_data
                p = cd['view_camera']
                data_time_start = cd['data_time_start']
                data_time_end = cd['data_time_end']
                time_interval = cd['data_time_interval']
                l= cam.objects.get(pk = p)
                c = inicialize()
                url_file = 'e:\\ftp\detektor\kassa\%s_MD.syslog' % (l.file_name)
                result = c.loadfile(url_file)
                vhod_detektor2 = 'event'+ ' ' + l.detektor3
                vihod_detektor2 = 'event'+' ' + l.detektor4
                vhod_detektor1 = 'event'+' ' + l.detektor1
                vihod_detektor1 = 'event'+ ' ' + l.detektor2
                ishod_arry = c.baza(result,vhod_detektor2,vihod_detektor2,vhod_detektor1,vihod_detektor1)
                c.zapis_v_basu(ishod_arry,l)
                variant = cd['vari_otchet']
                c.resultat = ['','','']
                init = False
                vall_vhod =[]
                if variant == '1':
                    init = True
                    data = event.objects.filter(camera_id = p, data_time__gte = data_time_start, data_time__lte = data_time_end)
                    c.interpritacia(data,l)
                nv = None
                if variant == '2':
                    data_count_day = data_time_end.day-data_time_start.day
                    n = 0
                    while n<data_count_day+1:
                        data = event.objects.filter(camera_id = p, data_time__gte = data_time_start + datetime.timedelta(days = n), data_time__lte = data_time_start+datetime.timedelta(days = n, hours=12))
                        c.interpritacia(data,l)
                        data_otbora = (data_time_start + datetime.timedelta(days = n)).date()
                        if not otchet_po_day.objects.filter(data = data_otbora, camera_ot = p ):
                            v = otchet_po_day.objects.create(camera_ot_id = p, data = data_otbora, vhod = c.resultat[0], vihod = c.resultat[1],srednee = c.resultat[2])
                            v.save()
                        else:
                            v = otchet_po_day.objects.filter(data = data_otbora, camera_ot = p)
                            for m in v:
                                if m.vhod < c.resultat[0]:
                                    m.vhod = c.resultat[0]
                                    m.save()
                                if m.vihod < c.resultat[1]:
                                    m.vihod = c.resultat[1]
                                    m.save()
                                if m.srednee < c.resultat[2]:
                                    m.srednee = c.resultat[2]
                                    m.save()
                        n+=1
                    nv = otchet_po_day.objects.filter(camera_ot_id = p, data__gte = data_time_start.date(), data__lte = data_time_end.date()).order_by('data')
                    vall_vhod = [[n.data, n.vhod, n.vihod, n.srednee] for n in nv]

                days = []
                den = None
                if variant == '3':
                    datas = otchet_po_hears.objects.all()
                    datas.delete()
                    data1 = None
                    data_time_start1 = data_time_start
                    while data_time_start < data_time_end:
                        minuts = time_interval.minute
                        hers = time_interval.hour
                        second = time_interval.second
                        data_time_start += datetime.timedelta(hours=hers, minutes = minuts, seconds= second)
                        data = event.objects.filter(camera_id = p, data_time__gte = data_time_start-datetime.timedelta(hours=hers, minutes=minuts, seconds=second), data_time__lte = data_time_start)
                        if not data:
                            continue
                        else:
                            if not otchet_po_hears.objects.filter(data_ot_day = data_time_start.date(), time_ot = data_time_start.time()):
                                c.interpritacia(data,l)
                                data1 = otchet_po_hears.objects.create(cam_ot_id = p, data_ot_day = data_time_start.date(), time_ot = data_time_start.time(), vhod_ot = c.resultat[0], vihod_ot = c.resultat[1], srednee_ot = c.resultat[2])
                                data1.save()
                    delta_den = data_time_end.day - data_time_start1.day
                    n = 0
                    days = []
                    vremen = []
                    while n<delta_den+1:
                        days.append(data_time_start1+datetime.timedelta(days = n))
                        vremen.append(otchet_po_hears.objects.filter(cam_ot_id = p, data_ot_day =(data_time_start1+datetime.timedelta(days = n)).date()).count())
                        n+=1
                    a = max(vremen)
                    b = vremen.index(a)
                    den = data_time_start1 + datetime.timedelta(days = b)
                return render_to_response('chitalka.html',{'form1':form1,'l':l, 'vhod':c.resultat[0],'vihod':c.resultat[1],'srednee':c.resultat[2],'otchet_po_day1':nv,'vall_vhod':vall_vhod,'days':days,'den':den,'init':init},RequestContext(request))
        elif request.method == 'POST' and request.POST.has_key('Zapis'):
            from openpyxl.workbook import Workbook
            from openpyxl.styles import Color, PatternFill, Font, Border, colors, Alignment, Side
            wb = Workbook()
            dest_filename = 'empty_book.xlsx'
            ws1 = wb.active
            ws1.title = "range names"
            url = None
            if variant == '3':
                delta_den = data_time_end.day - data_time_start1.day
                n = 0
                days = []
                vremen = []
                while n<delta_den+1:
                    days.append((data_time_start1+datetime.timedelta(days = n)).date())
                    vremen.append(otchet_po_hears.objects.filter(cam_ot_id = p, data_ot_day =(data_time_start1+datetime.timedelta(days = n)).date()).count())
                    n+=1
                a = max(vremen)
                b = vremen.index(a)
                den = data_time_start1 + datetime.timedelta(days = int(b))
                data = otchet_po_hears.objects.filter(data_ot_day = den.date(), cam_ot_id = p)
                ws1.merge_cells('A1:AB1')
                ws1['A1'] = u'Отчёт по дням в разрезе времени'
                ft = Font(name='Calibri', size=14, bold= True)
                al = Alignment(horizontal='center')
                a = ws1['A1']
                a.font = ft
                a.alignment = al
                ws1['A2'] = u'Дата'
                ft_data = Font(name='Calibri', size=12, bold= True)
                Al_data = Alignment(horizontal='center', vertical='center',text_rotation=0)
                B_data = Border(right=Side(border_style='thin'),
                                left=Side(border_style='thin'),
                                top=Side(border_style='thin'),
                                bottom=Side(border_style='thin'),
                                outline=Side(border_style='thin'))
                a = ws1['A2']
                a.font = ft_data
                a.alignment = Al_data
                a.border = B_data
                ws1['B2'] = u'Время'
                Al_time = Alignment(horizontal='center',text_rotation=90)
                ft_time = Font(name='Calibri', size=12, bold= True)
                a = ws1['B2']
                a.font = ft_time
                a.alignment = Al_time
                a.border = B_data
                i=1
                for rec in days:
                    ws1.merge_cells(start_row=i*3,start_column=1,end_row=(i*3)+2,end_column=1)
                    ws1.cell(row=i*3, column=1).value = rec
                    a = ws1.cell(row=i*3, column=1)
                    a.font = ft_data
                    a.alignment = Al_data
                    a.border = B_data
                    i+= 1
                i=3
                u=0
                for rec in data:
                    ws1.cell(row = 2, column = i).value = rec.time_ot
                    a = ws1.cell(row = 2, column = i)
                    a.font = ft_time
                    a.alignment = Al_time
                    a.border = B_data
                    i+=1
                    u = i
                i = 0
                while i < len(vremen)*3:
                    ws1.cell(row = i+3, column = 2).value = u'Вход'
                    ws1.cell(row = i+4, column = 2).value = u'Выход'
                    ws1.cell(row = i+5, column = 2).value = u'Среднее'
                    a1 = ws1.cell(row = i+3, column = 2)
                    a2 = ws1.cell(row = i+4, column = 2)
                    a3 = ws1.cell(row = i+5, column = 2)
                    a1.border,a2.border,a3.border = B_data,B_data,B_data
                    i+=3
                t = 0
                for rec in days:
                    data = otchet_po_hears.objects.filter(data_ot_day = rec)
                    i = 3
                    for m in data:
                        ws1.cell(row = t+3, column = i).value = m.vhod_ot
                        ws1.cell(row = t+4, column = i).value = m.vihod_ot
                        ws1.cell(row = t+5, column = i).value = m.srednee_ot
                        a1 = ws1.cell(row = t+3, column = i)
                        a2 = ws1.cell(row = t+4, column = i)
                        a3 = ws1.cell(row = t+5, column = i)
                        a1.font = ft_data
                        a2.font = ft_data
                        a3.font = ft_data
                        a1.border,a2.border,a3.border = B_data,B_data,B_data
                        i+=1
                    t+=3
                q =3
                while q<(len(vremen)*3)+3:
                    w =3
                    while w < u:
                        if not ws1.cell(row = q, column = w).value:
                            ws1.cell(row = q, column = w).value = 0
                            a = ws1.cell(row = q, column = w)
                            a.border = B_data
                        w+=1
                    q+=1
                wb.save(filename= 'c:\www\mysite\statics\%s' % (dest_filename))
                url = '/statics/empty_book.xlsx'
                p = u'Всё получилось, теперь можно скачать отчёт и поглядеть'
            if variant == '2':
                ws1.merge_cells('A1:AB1')
                ws1['A1'] = u'Отчёт по дням'
                ft = Font(name='Calibri', size=14, bold= True)
                al = Alignment(horizontal='center')
                aa = ws1['A1']
                aa.font = ft
                aa.alignment = al
                shapka = ['Дата', 'Вход', 'Выход', 'Среднее']
                B_data = Border(right=Side(border_style='thin'),
                                    left=Side(border_style='thin'),
                                    top=Side(border_style='thin'),
                                    bottom=Side(border_style='thin'),
                                    outline=Side(border_style='thin'))
                Al_data = Alignment(horizontal='center', vertical='center',text_rotation=0)
                ft_data = Font(name='Calibri', size=12, bold= True)
                data = otchet_po_day.objects.filter(camera_ot_id = p, data__gte = data_time_start.date(), data__lte = data_time_end.date())
                n = 2
                for mn in shapka:
                    ws1.cell(row = n, column = 1).value = mn
                    a = ws1.cell(row = n, column = 1)
                    a.font = ft_data
                    a.alignment = Al_data
                    a.border = B_data
                    n+=1
                n=2
                for m in data:
                    ws1.cell(row = 2, column = n).value = m.data
                    ws1.cell(row = 3, column = n).value = m.vhod
                    ws1.cell(row = 4, column = n).value = m.vihod
                    ws1.cell(row = 5, column = n).value = m.srednee
                    a1 = ws1.cell(row = 2, column = n)
                    a2 = ws1.cell(row = 3, column = n)
                    a3 = ws1.cell(row = 4, column = n)
                    a4 = ws1.cell(row = 5, column = n)
                    a1.font = ft_data
                    a1.alignment,a2.alignment,a3.alignment,a4.alignment = Al_data,Al_data,Al_data,Al_data
                    a1.border,a2.border,a3.border,a4.border = B_data,B_data,B_data,B_data
                    n+=1
                wb.save(filename= 'c:\www\mysite\statics\%s' % (dest_filename))
                url = '/statics/empty_book.xlsx'
                p=u'Все получилось, сохранилось'
            return  render_to_response('thanks.html',{'c':p,'url':url},RequestContext(request))
        else:
            form1 = view_cam()
        return render_to_response('chitalka.html',{'form1':form1},RequestContext(request))


class inicialize():
    def loadfile(self,vol=''):
        try:
            self.f = open(vol,'r')
        except IOError:
            self.error = u'Ощибка файла %s нет на диске' % vol
            return self.error
        self.ishod_file = []
        for self.n in self.f:
            self.ishod_file.append(self.n[:-1])
        self.f.close()
        return self.ishod_file
    def baza(self,result,vol1,vol2,vol3,vol4):
        self.itog_masiv = []
        self.vhod2 = 0
        for self.n in result:
            self.vhod2 = self.n.find(vol1)
            if self.vhod2 == -1:
                self.vhod2 = self.n.find(vol2)
                if self.vhod2 == -1:
                     self.vhod2 = self.n.find(vol3)
                     if self.vhod2 == -1:
                         self.vhod2 = self.n.find(vol4)
                         if self.vhod2 == -1:
                             continue
            self.now_time = datetime.datetime.now()
            self.god = self.now_time.year
            self.a = str(self.god)+' '+str(self.n[0:15])
            self.data_time = datetime.datetime.strptime(self.a, '%Y %b  %d %H:%M:%S')
            self.itog_masiv.append(self.data_time)
            self.itog_masiv.append(self.n[self.vhod2+6:])
        return self.itog_masiv

    def zapis_v_basu(self,ishod_arry, vol):
        self.total = event.objects.filter(camera_id = vol.id).count()
        if self.total>1:
            self.LastTernR = event.objects.filter(camera_id = vol.id)[self.total-1:self.total]
        else:
            self.LastTernR = event.objects.filter(camera_id = vol.id)[0:1]
        if not self.LastTernR:
            self.n=0
            while self.n<len(ishod_arry):
                self.q =event.objects.create(camera_id = vol.id, data_time = ishod_arry[self.n], events = ishod_arry[self.n+1])
                self.q.save()
                self.n+=2
        else:
            self.n = 0
            self.a = self.LastTernR.get().data_time.strftime('%Y %b  %d %H:%M:%S')
            self.a = datetime.datetime.strptime(self.a, '%Y %b  %d %H:%M:%S')
            self.b = ishod_arry[len(ishod_arry)-2]
#            self.b = datetime.datetime.strftime(self.b, '%Y %b  %d %H:%M:%S')
            self.c = ishod_arry[0]
#            self.c = datetime.datetime.strftime(self.c, '%Y %b  %d %H:%M:%S')
            if self.a <= self.b:
                while self.n<len(ishod_arry):
                    self.b = ishod_arry[self.n]
#                    self.b = datetime.datetime.strftime(self.b,'%Y %b  %d %H:%M:%S')
                    if self.a == self.b:
                        ishod_arry = ishod_arry[self.n:]
                        ishod_arry = ishod_arry[2:]
                        self.t = 0
                        while self.t<len(ishod_arry):
                            self.q =event.objects.create(camera_id = vol.id, data_time = ishod_arry[self.t], events = ishod_arry[self.t+1])
                            self.q.save()
                            self.t+=2
                    self.n+=2
            if self.a < self.c:
                self.n = 0
                while self.n<len(ishod_arry):
#                    self.b = ishod_arry[self.n]
#                    self.b = datetime.datetime.strftime(self.b,'%Y %b  %d %H:%M:%S')
                    self.q =event.objects.create(camera_id = vol.id, data_time = ishod_arry[self.n], events = ishod_arry[self.n+1])
                    self.q.save()
                    self.n+=2
            else:
                return "asas"
        return self
    def interpritacia(self,data, data1):
        self.arry = []
        for self.n in data:
            self.arry.append(self.n.data_time)
            self.arry.append(self.n.events)
        self.n = 1
        self.count_vhod = 0
        self.count_vihod = 0
        self.resultat = []
        while self.n<len(self.arry)-6:
            if self.arry[self.n] == data1.detektor1 and self.arry[self.n+2] == data1.detektor2 and self.arry[self.n+4] == data1.detektor3 and self.arry[self.n+6] == data1.detektor4:
#                print u"вход"
                self.count_vhod +=1
            if self.arry[self.n] == data1.detektor3 and self.arry[self.n+2] == data1.detektor4 and self.arry[self.n+4] == data1.detektor1 and self.arry[self.n+6] == data1.detektor2:
#                print u'Выход'
                self.count_vihod+=1
            self.n+=2
        self.resultat.append(self.count_vhod)
        self.resultat.append(self.count_vihod)
        self.resultat.append((int(self.count_vhod)+int(self.count_vihod))/2)
        return self.resultat

def risovalka(request):
    return render_to_response('risov.html',{'request':request})
def isoval(request):
    if request.method == 'POST':
        data1 = request.POST['data']
        print data1

    return render_to_response('risov.html',{'request':request})

#Новый проект расчета непонятный
@login_required()
def settings_raschet(request,vol):
    data_Name = Seria_name.objects.all()
    data_Seria = serega.objects.filter(seria_name_id = vol)
    form_serg = Name_seria(request.POST)
    form_serg1 = Seria(request.POST)
    if request.method =='POST' and request.POST.has_key('add'):
        if form_serg.is_valid() or form_serg1.is_valid():
            cd1 = form_serg.cleaned_data
            name1 = cd1['name']
            s = Seria_name.objects.create(name = name1)
            s.save()
            return render_to_response('serega_add.html',{'form_serg':form_serg,'data':data_Name,'data1':data_Seria},RequestContext(request))
    if request.method == 'POST' and request.POST.has_key('Delete'):
        a = request.POST.getlist('znac')
        for n in a:
            a = Seria_name.objects.get(pk = n)
            b = serega.objects.filter(seria_name_id = n)
            a.delete()
            b.delete()
    if request.method == 'POST' and request.POST.has_key('insert'):
        if form_serg.is_valid() or form_serg1.is_valid():
            cd1 = form_serg.cleaned_data
            a = request.POST.getlist('znac')
            name = cd1['name']
            baz = Seria_name.objects.get(pk = a[0])
            baz.name = name
            baz.save()
            return render_to_response('serega_add.html',{'form_serg':form_serg,'data':data_Name,'data1':data_Seria},RequestContext(request))
    if request.method == 'POST' and request.POST.has_key('add1'):
        if  form_serg.is_valid() or form_serg1.is_valid():
            name_a = request.POST.getlist('name')
            name_b = request.POST.getlist('seria')
            s = serega.objects.create(seria_name_id = vol, name = name_a[0], seria = name_b[0])
            s.save()
            return render_to_response('serega_add.html',{'form_serg1':form_serg1,'data':data_Name,'data1':data_Seria},RequestContext(request))
    if request.method == 'POST' and request.POST.has_key('Delete1'):
        a = request.POST.getlist('znac1')
        for n in a:
            b = serega.objects.get(pk = n)
            b.delete()
    if request.method == 'POST' and request.POST.has_key('insert1'):
        if  form_serg.is_valid() or form_serg1.is_valid():
            a = request.POST.getlist('znac1')
            name_a = request.POST.getlist('name')
            name_b = request.POST.getlist('seria')
            baz = serega.objects.get(pk = a[0])
            baz.name = name_a[0]
            baz.seria = name_b[0]
            baz.save()
            return render_to_response('serega_add.html',{'form_serg1':form_serg1,'data':data_Name,'data1':data_Seria},RequestContext(request))
    else:
        if vol == '0':
            selectseria = False
        else:
            selectseria = Seria_name.objects.get(pk = vol)
        form_serg = Name_seria()
        form_serg1 = Seria()
    return render_to_response('serega_add.html',{'form_serg':form_serg,'form_serg1':form_serg1,'data':data_Name,'data1':data_Seria,'seria1':selectseria},RequestContext(request))



def raschet(request):
    global mas_ch, mas_ch_obr
    if request.method == 'POST':
        form2 = raschet_form(request.POST)
        if form2.is_valid():
            cd1 = form2.cleaned_data
            ch = cd1['chislo']
            ch1 = cd1['chislo1']
            ch2 = cd1['chislo2']
            ser = cd1['serga_vari']
            if ch == u'0' or ch == u'':
                ch = 1.0
            elif ch1 == u'0' or ch1 == u'':
                ch1 = 1.0
            elif ch2 == u'0' or ch2 == u'':
                ch2 = 1.0
            mas_ch = []
            if ser == u'0':
                mas_ch.append(ch)
                mas_ch.append(ch1)
                mas_ch.append(ch2)
            else:
                ser_a = serega.objects.filter(seria_name_id = ser)
                ff = []
                for n in ser_a:
                    ff.append(n.seria)
                mas_ch.append(ch)
                mas_ch += ff
            mas_ch_obr = [1/float(x) for x in mas_ch]
            d = float(ch)
            a = func_rasceta(d)
            it_mass1,it_mass,it_mass2,it_mass3,it_mass4,it_mass5,it_mass6,it_mass7 = [],[],[],[],[],[],[],[]
            it_mass12,it_mass_2,it_mass22,it_mass32 = [],[],[],[]
            it_mass13,it_mass_3,it_mass23,it_mass33 = [],[],[],[]
            it_mass14,it_mass24 = [],[]
            m_ch = 1
            while m_ch<len(mas_ch):
# Первый модуль
                a_chislo = func_rasceta(float(mas_ch[m_ch]))
                a_1group = a[1:5]
                a_chislo1_1g = a_chislo[1:5]
                sd = poisk_sovpad(a_1group,a_chislo1_1g)
                n = 2
                mass1,mass,mass2,mass3,mass4,mass5,mass6,mass7 = [],[],[],[],[],[],[],[]
                while n<len(sd):
                    gg = uslovie(a,sd,n,0,mas_ch[m_ch])
                    if sd[n] == 0 or sd[n]==1 or sd[n]==2 or sd[n]==3:
                        a[0][sd[n-2]] = str(a[0][sd[n-2]])
                    if sd[n] == 0:
                        a[1][sd[n-2]] = str(a[1][sd[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass1.append(gg[0])
                            mass.append(gg[1])
                    if sd[n] == 1:
                        a[2][sd[n-2]] = str(a[2][sd[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass2.append(gg[0])
                            mass3.append(gg[1])
                    if sd[n] == 2:
                        a[3][sd[n-2]] = str(a[3][sd[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass4.append(gg[0])
                            mass5.append(gg[1])
                    if sd[n] == 3:
                        a[4][sd[n-2]] = str(a[4][sd[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass6.append(gg[0])
                            mass7.append(gg[1])
                    n+=5
                it_mass1 += mass1
                it_mass += mass
                it_mass2 += mass2
                it_mass3 += mass3
                it_mass4 += mass4
                it_mass5 += mass5
                it_mass6 += mass6
                it_mass7 += mass7
# Второй модуль
                a_2group = a[7:9]
                a_chislo1_2g = a_chislo[7:9]
                sd1 = poisk_sovpad(a_2group,a_chislo1_2g)
                n = 2
                mass12,mass_2,mass22,mass32 = [],[],[],[]
                while n<len(sd1):
                    gg = uslovie1(a,sd1,n,7,mas_ch[m_ch])
                    if sd1[n] == 0 or sd1[n]==1 or sd1[n]==2 or sd1[n]==3:
                        a[6][sd1[n-2]] = str(a[6][sd1[n-2]])
                    if sd1[n] == 0:
                        a[7][sd1[n-2]] = str(a[7][sd1[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass12.append(gg[0])
                            mass_2.append(gg[1])
                    if sd1[n] == 1:
                        a[8][sd1[n-2]] = str(a[8][sd1[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass22.append(gg[0])
                            mass32.append(gg[1])

                    n +=5
                it_mass12 += mass12
                it_mass_2 += mass_2
                it_mass22 += mass22
                it_mass32 += mass32
# Третий модуль
                a_3group = a[10:12]
                a_chislo1_3g = a_chislo[10:12]
                sd2 = poisk_sovpad(a_3group,a_chislo1_3g)
                n = 2
                mass13,mass_3,mass23,mass33 = [],[],[],[]
                while n<len(sd2):
                    gg = uslovie2(a,sd2,n,10,mas_ch[m_ch])
                    if sd2[n] == 0 or sd2[n]==1 or sd2[n]==2 or sd2[n]==3:
                        a[9][sd2[n-2]] = str(a[9][sd2[n-2]])
                    if sd2[n] == 0:
                        a[10][sd2[n-2]] = str(a[10][sd2[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass13.append(gg[0])
                            mass_3.append(gg[1])
                    if sd2[n] == 1:
                        a[11][sd2[n-2]] = str(a[11][sd2[n-2]])
                        if gg[0] != 0 and gg[1] != 0:
                            mass23.append(gg[0])
                            mass33.append(gg[1])
                    n +=5
                it_mass13 += mass13
                it_mass_3 += mass_3
                it_mass23 += mass23
                it_mass33 += mass33
#Четвёртый модуль
                a_4group = a[37]
                a_chislo1_4g = a_chislo[37]
                sd3 = poisk_sovpad(a_4group,a_chislo1_4g)
                n=2
                mass14,mass24 = [],[]
                while n<len(sd3):
                    ggz = uslovie3(a,sd3,n,37,mas_ch[m_ch])
                    a[37][sd3[n]][sd3[n-2]] = str(a[37][sd3[n]][sd3[n-2]])
                    if ggz[0] != 0 and ggz[1] != 0:
                        mass14.append(ggz[0])
                        mass24.append(ggz[1])
                        n+=5
                    else:
                        n+=5
                it_mass14 += mass14
                it_mass24 += mass24
                m_ch+=1
            return render_to_response('raschet.html',{'form2':form2,'a':a[0],'a1':a[1],'a2':a[2],'a3':a[3],'a4':a[4],'a5':a[5],'a6':a[6],'a7':a[7],
                                            'a8':a[8],'a9':a[9],'a10':a[10],
                                            'obrat':1/d,'den':a[12],'nedel':a[13],'moth':a[14],'let':a[15],'hear':a[16],'minuts':a[17],'seconds':a[18],
                                            'gipot':a[19],'gipot1':a[20],'gipot2':a[21],'gipot3':a[22],'kub':a[23],'kub1':a[24],'kub2':a[25],'kub3':a[26],'kub4':a[27],'kub5':a[28],
                                            'kub6':a[29],'kubb':a[30],'kubb1':a[31],'kubb2':a[32],'obrat_360':a[11], 'mass':it_mass,'mass1':it_mass1,'mass2':it_mass2,'mass3':it_mass3,'mass4':it_mass4,'mass5':it_mass5,
                                            'mass6':it_mass6,'mass7':it_mass7, 'mas_ch':mas_ch,'mas_ch_obr':mas_ch_obr, 'it_mass12':it_mass12,'it_mass2':it_mass_2,
                                            'it_mass22':it_mass22,'it_mass32':it_mass32,'it_mass13':it_mass13,'it_mass_3':it_mass_3,
                                            'it_mass23':it_mass23,'it_mass33':it_mass33,'it_mass14':it_mass14,'it_mass24':it_mass24,'figura_coment':a[33],'figura_racsceta':a[34],'figured_rasceta_obr':a[35],'dni1':a[36],'dni2':a[37]}, RequestContext(request))
    else:
        form2 = raschet_form()
    return render_to_response('raschet.html',{'form2':form2}, RequestContext(request))

def raschet1(request, znach):
    znach = str(znach)
    znach_a = znach.split(',')
    znach = '.'.join(znach_a)
    try:
        znach = float(znach)
    except ValueError:
        raise Http404()
    d= float(znach)
    a = func_rasceta(d)
    it_mass1,it_mass,it_mass2,it_mass3,it_mass4,it_mass5,it_mass6,it_mass7 = [],[],[],[],[],[],[],[]
    it_mass12,it_mass_2,it_mass22,it_mass32 = [],[],[],[]
    it_mass13,it_mass_3,it_mass23,it_mass33 = [],[],[],[]
    it_mass14,it_mass24 = [],[]
    m_ch = 1
    while m_ch<len(mas_ch):
# Первый модуль
        a_chislo = func_rasceta(float(mas_ch[m_ch]))
        a_1group = a[1:5]
        a_chislo1_1g = a_chislo[1:5]
        sd = poisk_sovpad(a_1group,a_chislo1_1g)
        n = 2
        mass1,mass,mass2,mass3,mass4,mass5,mass6,mass7 = [],[],[],[],[],[],[],[]
        while n<len(sd):
            gg = uslovie(a,sd,n,0,mas_ch[m_ch])
            if sd[n] == 0 or sd[n]==1 or sd[n]==2 or sd[n]==3:
                a[0][sd[n-2]] = str(a[0][sd[n-2]])
            if sd[n] == 0:
                a[1][sd[n-2]] = str(a[1][sd[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass1.append(gg[0])
                    mass.append(gg[1])
            if sd[n] == 1:
                a[2][sd[n-2]] = str(a[2][sd[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass2.append(gg[0])
                    mass3.append(gg[1])
            if sd[n] == 2:
                a[3][sd[n-2]] = str(a[3][sd[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass4.append(gg[0])
                    mass5.append(gg[1])
            if sd[n] == 3:
                a[4][sd[n-2]] = str(a[4][sd[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass6.append(gg[0])
                    mass7.append(gg[1])
            n+=5
        it_mass1 += mass1
        it_mass += mass
        it_mass2 += mass2
        it_mass3 += mass3
        it_mass4 += mass4
        it_mass5 += mass5
        it_mass6 += mass6
        it_mass7 += mass7
#Второй модуль
        a_2group = a[7:9]
        a_chislo1_2g = a_chislo[7:9]
        sd1 = poisk_sovpad(a_2group,a_chislo1_2g)
        n = 2
        mass12,mass_2,mass22,mass32 = [],[],[],[]
        while n<len(sd1):
            gg = uslovie1(a,sd1,n,7,mas_ch[m_ch])
            if sd1[n] == 0 or sd1[n]==1 or sd1[n]==2 or sd1[n]==3:
                a[6][sd1[n-2]] = str(a[6][sd1[n-2]])
            if sd1[n] == 0:
                a[7][sd1[n-2]] = str(a[7][sd1[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass12.append(gg[0])
                    mass_2.append(gg[1])
            if sd1[n] == 1:
                a[8][sd1[n-2]] = str(a[8][sd1[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass22.append(gg[0])
                    mass32.append(gg[1])
            n +=5
        it_mass12 += mass12
        it_mass_2 += mass_2
        it_mass22 += mass22
        it_mass32 += mass32
#Третий модуль
        a_3group = a[10:12]
        a_chislo1_3g = a_chislo[10:12]
        sd2 = poisk_sovpad(a_3group,a_chislo1_3g)
        n = 2
        mass13,mass_3,mass23,mass33 = [],[],[],[]
        while n<len(sd2):
            gg = uslovie2(a,sd2,n,10,mas_ch[m_ch])
            if sd2[n] == 0 or sd2[n]==1 or sd2[n]==2 or sd2[n]==3:
                a[9][sd2[n-2]] = str(a[9][sd2[n-2]])
            if sd2[n] == 0:
                a[10][sd2[n-2]] = str(a[10][sd2[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass13.append(gg[0])
                    mass_3.append(gg[1])
            if sd2[n] == 1:
                a[11][sd2[n-2]] = str(a[11][sd2[n-2]])
                if gg[0] != 0 and gg[1] != 0:
                    mass23.append(gg[0])
                    mass33.append(gg[1])
            n +=5
        it_mass13 += mass13
        it_mass_3 += mass_3
        it_mass23 += mass23
        it_mass33 += mass33
#Четвёртый модуль
        a_4group = a[37]
        a_chislo1_4g = a_chislo[37]
        sd3 = poisk_sovpad(a_4group,a_chislo1_4g)
        n=2
        mass14,mass24 = [],[]
        while n<len(sd3):
            ggz = uslovie3(a,sd3,n,37,mas_ch[m_ch])
            a[37][sd3[n]][sd3[n-2]] = str(a[37][sd3[n]][sd3[n-2]])
            mass14.append(ggz[0])
            mass24.append(ggz[1])
            n+=5
        it_mass14 += mass14
        it_mass24 += mass24
        m_ch += 1
    return render_to_response('raschet1.html',{'a':a[0],'a1':a[1],'a2':a[2],'a3':a[3],'a4':a[4],'a5':a[5],'a6':a[6],'a7':a[7],
                                    'a8':a[8],'a9':a[9],'a10':a[10],
                                    'obrat':1/d,'den':a[12],'nedel':a[13],'moth':a[14],'let':a[15],'hear':a[16],'minuts':a[17],'seconds':a[18],
                                    'gipot':a[19],'gipot1':a[20],'gipot2':a[21],'gipot3':a[22],'kub':a[23],'kub1':a[24],'kub2':a[25],'kub3':a[26],'kub4':a[27],'kub5':a[28],
                                    'kub6':a[29],'kubb':a[30],'kubb1':a[31],'kubb2':a[32],'obrat_360':a[11], 'mass':it_mass,'mass1':it_mass1,'mass2':it_mass2,'mass3':it_mass3,'mass4':it_mass4,'mass5':it_mass5,
                                    'mass6':it_mass6,'mass7':it_mass7, 'mas_ch':mas_ch,'mas_ch_obr':mas_ch_obr,'it_mass12':it_mass12,'it_mass2':it_mass_2,
                                    'it_mass22':it_mass22,'it_mass32':it_mass32,'it_mass13':it_mass13,'it_mass_3':it_mass_3,
                                    'it_mass23':it_mass23,'it_mass33':it_mass33,'it_mass14':it_mass14,'it_mass24':it_mass24,'figura_coment':a[33],'figura_racsceta':a[34],'figured_rasceta_obr':a[35],'dni1':a[36],'dni2':a[37]}, RequestContext(request))


def func_rasceta(val):
# начало
    chisl = [1,2,3,5,6,7,12]
    leght = [2,16,9,4,10,9,4]
    itog,fff =[],[]
    ra = [2,3,5,7,10,12,'Пи','Кор2+1','Кор2-1','Кор2(Кор2+1)','Кор2(Кор2-1)','Пи^2','Пи^3','Пи^4','Пи^5','Пи^6','Кор Пи','фи','Фи','Фи2','Кор Фи','Число Е','Е в квадрате','Пи-E']
    ramath = [stepen(2,0.5),stepen(3,0.5),stepen(5,0.5),stepen(7,0.5),stepen(10,0.5),stepen(12,0.5),math.pi,stepen(2,0.5)+1,stepen(2,0.5)-1,stepen(2,0.5)*(stepen(2,0.5)+1),stepen(2,0.5)*(stepen(2,0.5)-1),stepen(math.pi,2),stepen(math.pi,3),stepen(math.pi,4),stepen(math.pi,5),stepen(math.pi,6),stepen(math.pi,0.5),0.6180339887,1.6180339887,2.6180339887,stepen(1.6180339887,0.5),math.e,stepen(math.e,2),math.pi-math.e]
    palneta_cons = [27.32,365.25,88,224.7,686.97,4332.589,10759.22,30685.4,60190.03,90553.02,6585.033,6798.3835]
    nn = 0
    while nn<len(chisl):
        res = [chisl[nn]*stepen(chisl[nn],x-1) for x in xrange(1,leght[nn])]
        itog += res
        nn+=1
    ra = itog + ra
    itog += ramath
    itog += palneta_cons
    ggg = []
    ggg1 = []
    l=360
    for n in palneta_cons:
        ggg.append((val/n)*l)
        ggg1.append(1/((val/n)*l))
    planeta = ['луна','земля','меркурий','венера','марс','Юпитер','Сатурн','Уран','Нептун','Плутон','Сарос','Узел']
    ra += planeta

    nedel = val/7
    moth = val/30.4375
    let = val/365.25
    hear = val*24
    minuts = val*24*60
    seconds = val*24*60*60

    matrica_dnei_opis = ['Дни','Недели','Месяцы','Часы','Минуты','Секунды','Годы']
    matrica_dnei = [[val,val/7,val/30.4375,val*24,val*24*60,val*24*3600,val/365.25],[val*7,val,(val*7)/30.4375,val*7*24,val*7*24*60,val*7*24*3600,val*7/365.25],[val*30.4375,(val*30.4375)/7,val,val*30.4375*24,val*30.4375*24*60,val*30.4375*24*3600,val*30.4375/365.25],
                    [val/24,(val/24)/7,(val/24)/30.4375,val,val*60,val*3600,val/(24*365.25)],[(val/60)/24,val/(60*24*7),val/(60*24*30.4375),val/60,val,(val/60)*3600,val/(60*24*365.25)],
                    [(val/3600)/24,val/(3600*24*7),val/(3600*24*30.4375),val/3600,val/60,val,val/(3600*24*365.25)],[val*365.25,val*365.25/7,val*365.25/30.4375,val*365.25*24,val*365.25*24*60,val*365.25*24*3600,val]]

    multi_umn = []
    multi_delen = []
    multi_umn_obr = []
    multi_delen_obr = []
    for n in itog:
        multi_umn.append(val*n)
        multi_umn_obr.append(1/(val*n))
        multi_delen.append(val/n)
        multi_delen_obr.append(1/(val/n))
    stepeni = [2,3,4,5,6,7,0.5,0.33,0.25,0.2,0.166666666666667,0.142857142857143,7.665,1/7.665,9.5785,1/9.5785]
    multi_stepeni = []
    multi_stepeni_obr=[]
    for n in stepeni:
        multi_stepeni.append(stepen(val,n))
        multi_stepeni_obr.append(1/(stepen(val,n)))

    gipot = (2*(val/stepen(2,0.5)))+val
    gipot1 = val+val+(val*stepen(2,0.5))
    gipot2 = val/(stepen(2,0.5)+1)
    gipot3 = gipot2/stepen(2,0.5)
    kub = val*stepen(3,0.5)
    kub1 = kub/2
    kub2 = val*stepen(2,0.5)
    kub3 = val*12
    kub4 = val/stepen(2,0.5)
    kub5 = kub4*12
    kub6 = kub4*stepen(3,0.5)
    kubb = val/stepen(3,0.5)
    kubb1 = kubb*12
    kubb2 = kubb*stepen(2,0.5)
    figured_coments = ['Треугольник', 'Пер. прям тр. со стор. 1,1 кор2','Если изв гепот','Если изв сторон','Если изв перим треуг','то гипотинуза', 'то сторона','куб из грань','диогональ',\
                       'по диогонали','диаг грани','периметр грани','куб изв диаг грани','грань','периметр граней','диагональ','куб изв.диагон','грань','периметр','диагон. Грани']
    figured_rasceta = [111111,111111,gipot,gipot1,111111,gipot2,gipot3,111111,kub,kub1,kub2,kub3,111111,kub4,kub5,kub6,111111,kubb,kubb1,kubb2]
    figured_rasceta_obr = []
    for n in figured_rasceta:
        if n == 111111:
            figured_rasceta_obr.append(n)
        else:
            figured_rasceta_obr.append(1/n)
    return itog,multi_umn,multi_delen,multi_umn_obr,multi_delen_obr,ra,\
           stepeni,multi_stepeni,multi_stepeni_obr,planeta, ggg,ggg1,val,nedel,moth,let,hear,minuts,seconds,gipot,gipot1,gipot2,gipot3,\
           kub,kub1,kub2,kub3,kub4,kub5,kub6,kubb,kubb1,kubb2,figured_coments,figured_rasceta,figured_rasceta_obr,matrica_dnei_opis,matrica_dnei

def stepen(x,y):
     return math.pow(x,y)

def poisk_sovpad(vol,vol1):
    i = 0
    vot = []
    kol1=0
    while i<len(vol):
        if type(vol[i]) == list:
            kol = 0
            while kol<len(vol1):
                j = 0
                for n in vol[i]:
                    k=0
                    for m in vol1[kol]:
                        if razbor_masiva(n) == razbor_masiva(m):
#                            print razbor_masiva(n),'   ',razbor_masiva(m),'  ',n,'   ',m
                            vot.append(j)
                            vot.append(k)
                            vot.append(kol1)
                            vot.append(kol)
                            vot.append(razbor_masiva(m))
                            k+=1
                        k+=1
                    j+=1
                kol+=1
        else:
            q = 0
            while q<len(vol1):
                if type(vol1[q]) == list:
                    qq = 0
                    for n in vol1[q]:
                        if razbor_masiva(vol[i]) == razbor_masiva(vol1[n]):
                            vot.append(i)
                            vot.append(qq)
                            vot.append(q)
                            vot.append(razbor_masiva(vol1[n]))
                            qq+=1
                        qq+=1
                else:
                    if razbor_masiva(vol[i]) == razbor_masiva(vol1[q]):
                        vot.append(i)
                        vot.append(q)
                        vot.append(razbor_masiva(vol1[q]))
        i+=1
        kol1+=1
    return vot

def razbor_masiva(vol):
    h = str(vol)
    hh = h.split('.')
    hh = ''.join(hh)
    n = 0
    k = 0
    while n<len(hh):
        if hh[n] == '0':
            k+=1
        else:
            break
        n+=1
    hh = hh[k:]
    if hh[0] !='0':
        if len(hh)<=5:
            rez = hh[:len(hh)]
        else:
            m=0
            while m<5:
                if hh[m]== 'e':
                    break
                m+=1
            rez = hh[:m]
    else:
        rez = hh[:6]
#    rez = int(rez)
    return rez

def uslovie(vol,vol1,vol2,vol3,vol4): #'''vol = a, vol1 = sd, vol2 = n'''
    qz = serega.objects.get(seria = vol4)
    if vol1[vol2] == 0:
        indexing = 1
    elif vol1[vol2] == 1:
        indexing  = 2
    elif vol1[vol2] == 2:
        indexing = 3
    elif vol1[vol2] == 3:
        indexing = 4
    if vol1[vol2+1] == 0:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '%s * %s = %s</br>' % (vol[0+vol3][vol1[vol2-1]],vol4,vol[indexing+vol3][vol1[vol2-2]])
    elif vol1[vol2+1] == 1:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '%s / %s = %s</br>' % (vol4,vol[0+vol3][vol1[vol2-1]],vol[indexing+vol3][vol1[vol2-2]])
    elif vol1[vol2+1] == 2:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '1/(%s * %s) = %s</br>' % (vol4,vol[0+vol3][vol1[vol2-1]],vol[indexing+vol3][vol1[vol2-2]])
    elif vol1[vol2+1] == 3:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '1/(%s / %s) = %s</br>' % (vol4,vol[0+vol3][vol1[vol2-1]],vol[indexing+vol3][vol1[vol2-2]])
    else:
        mass,mass1 = 0,0
    return mass1,mass

def uslovie1(vol,vol1,vol2,vol3,vol4): #'''vol = a, vol1 = sd, vol2 = n'''
    qz = serega.objects.get(seria = vol4)
    if vol1[vol2] == 0:
        indexing = 0
    elif vol1[vol2] == 1:
        indexing  = 1
    if vol1[vol2+1] == 0:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '%s^%s = %s</br>' % (vol4,vol[-1+vol3][vol1[vol2-1]],vol[indexing+vol3][vol1[vol2-2]])
    elif vol1[vol2+1] == 1:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '1/(%s^%s) = %s</br>' % (vol4,vol[-1+vol3][vol1[vol2-1]],vol[indexing+vol3][vol1[vol2-2]])
    else:
        mass,mass1 = 0,0
    return mass1,mass

def uslovie2(vol,vol1,vol2,vol3,vol4): #'''vol = a, vol1 = sd, vol2 = n'''
    qz = serega.objects.get(seria = vol4)
    if vol1[vol2] == 0:
        indexing = 0
    elif vol1[vol2] == 1:
        indexing  = 1
    if vol1[vol2+1] == 0:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '(%s/%s)*360 = %s</br>' % (vol4,vol[0][-12+vol1[vol2-1]:-12+vol1[vol2-1]+1],vol[indexing+vol3][vol1[vol2-2]])
    elif vol1[vol2+1] == 1:
        mass1 = '%s | %s' % (qz.name,vol[indexing+vol3][vol1[vol2-2]])
        mass = '1/(%s/%s)*360 = %s</br>' % (vol4,vol[0][-12+vol1[vol2-1]:-12+vol1[vol2-1]+1],vol[indexing+vol3][vol1[vol2-2]])
    else:
        mass,mass1 = 0,0
    return mass1,mass

def uslovie3(vol,vol1,vol2,vol3,vol4): #'''vol = a, vol1 = sd, vol2 = n'''
    qz = serega.objects.get(seria = vol4)
    mass1 = '%s | %s' % (qz.name,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    if vol1[vol2+1]==0 and vol1[vol2-1]==0:
        mass = '%s=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==1 and vol1[vol2-1]==0:
        mass = '%s*7=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==2 and vol1[vol2-1]==0:
        mass = '%s*30.4375=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==3 and vol1[vol2-1]==0:
        mass = '%s/24=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==4 and vol1[vol2-1]==0:
         mass = '%s/(60*24)=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==5 and vol1[vol2-1]==0:
        mass = '%s/(3600*24)=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==6 and vol1[vol2-1]==0:
        mass = '%s*365.25=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==0 and vol1[vol2-1]==1:
        mass = '%s/7=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==0 and vol1[vol2-1]==2:
        mass = '%s/30,4375=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==0 and vol1[vol2-1]==3:
        mass = '%s*24=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==0 and vol1[vol2-1]==4:
        mass = '%s*24*60=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==0 and vol1[vol2-1]==5:
        mass = '%s*24*3600=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==0 and vol1[vol2-1]==6:
        mass = '%s/365.25=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==1 and vol1[vol2-1]==1:
        mass = '%s=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==1 and vol1[vol2-1]==2:
        mass = '%s*7/30,4375=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==1 and vol1[vol2-1]==3:
        mass = '%s*7*24=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==1 and vol1[vol2-1]==4:
        mass = '%s*7*24*60=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==1 and vol1[vol2-1]==5:
        mass = '%s*7*24*3600=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==1 and vol1[vol2-1]==6:
        mass = '%s*7/365,25=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==2 and vol1[vol2-1]==1:
        mass = '%s*30,4375/7=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==2 and vol1[vol2-1]==2:
        mass = '%s=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==2 and vol1[vol2-1]==3:
        mass = '%s*30,4375*24=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==2 and vol1[vol2-1]==4:
        mass = '%s*30,4375*24*60=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==2 and vol1[vol2-1]==5:
        mass = '%s*30,4375*24*3600=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==2 and vol1[vol2-1]==6:
        mass = '%s*30.4375/365.25=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==3 and vol1[vol2-1]==1:
        mass = '%s/24/7=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==3 and vol1[vol2-1]==2:
        mass = '%s/24/30,4375=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==3 and vol1[vol2-1]==3:
        mass = '%s=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==3 and vol1[vol2-1]==4:
        mass = '%s*60=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==3 and vol1[vol2-1]==5:
        mass = '%s*3600=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==3 and vol1[vol2-1]==6:
        mass = '%s/(24*365,25)=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==4 and vol1[vol2-1]==1:
        mass = '%s/24/7=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==4 and vol1[vol2-1]==2:
        mass = '%s/24/30,4375=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==4 and vol1[vol2-1]==3:
        mass = '%s/60=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==4 and vol1[vol2-1]==4:
        mass = '%s=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==4 and vol1[vol2-1]==5:
        mass = '%s/60*3600=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==4 and vol1[vol2-1]==6:
        mass = '%s/(60*24*365,25)=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==5 and vol1[vol2-1]==1:
        mass = '%s/24/7=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==5 and vol1[vol2-1]==2:
        mass = '%s/24/30,4375=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==5 and vol1[vol2-1]==3:
        mass = '%s/3600=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==5 and vol1[vol2-1]==4:
        mass = '%s/60=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==5 and vol1[vol2-1]==5:
        mass = '%s=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==5 and vol1[vol2-1]==6:
        mass = '%s/(24*3600*365,25)=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    elif vol1[vol2+1]==6 and vol1[vol2-1]==6:
        mass = '%s=%s</br>' % (vol4,vol[vol3][vol1[vol2]][vol1[vol2-2]])
    else:
        mass,mass1 = 0,0
    return mass1,mass

